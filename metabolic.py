from openpyxl import load_workbook
import numpy as np
from keras.models import Sequential
from keras.utils import np_utils
from keras.layers.core import Dense, Activation
from random import shuffle

data = []
output = []
ws = load_workbook('../SAPBA_WP.xlsx').active
CIMT = 1
sick = 0
healthy = 0
index = 16 #11
for row in ws.iter_rows(min_row=2):
    print(row[4].value,row[5].value)
    if (type(row[index].value) == int and row[4].value<135.75 and row[4].value > 48.42 and row[5].value <170 and row[5].value>90):
        
        if row[index].value ==  CIMT: 
            while True:      
                data.append((row[3].value,row[4].value,row[5].value))
                output.append(1) 
                sick +=1
                if healthy <= sick:
                    break
        else:
            while True:
                data.append((row[3].value,row[4].value,row[5].value))
                output.append(0)
                healthy+=1
                if healthy >= sick:
                    break 
combined = list(zip(data, output))
shuffle(combined)
data[:], output[:] = zip(*combined)

x_train = np.array(data[:np.ceil(len(data)*0.9).astype('uint16')])
x_test = np.array(data[np.ceil(len(data)*0.9).astype('uint16'):])

y_train = np.array(output[:np.ceil(len(output)*0.9).astype('uint16')])
y_train = np_utils.to_categorical(y_train) 
y_test = np.array(output[np.ceil(len(output)*0.9).astype('uint16'):])
y_test = np_utils.to_categorical(y_test) 

scale = np.max(x_train)
x_train /= scale
x_test /= scale

mean = np.std(x_train)
x_train -= mean
x_test -= mean
    
input_dim = x_train.shape[1]
nb_classes = y_train.shape[1]

model = Sequential()
model.add(Dense(32, input_dim=input_dim))
model.add(Activation('relu'))
# model.add(Dense(16))
# model.add(Activation('relu'))
model.add(Dense(nb_classes))
model.add(Activation('softmax'))
model.compile(loss='binary_crossentropy', optimizer='adadelta',metrics=['accuracy'])

TP,FP,TN,FN = (0,0,0,0)
def shuffle_weights(model, weights=None):
    """Randomly permute the weights in `model`, or the given `weights`.
    This is a fast approximation of re-initializing the weights of a model.
    Assumes weights are distributed independently of the dimensions of the weight tensors
      (i.e., the weights have the same distribution along each dimension).
    :param Model model: Modify the weights of the given model.
    :param list(ndarray) weights: The model's weights will be replaced by a random permutation of these weights.
      If `None`, permute the model's current weights.
    """
    if weights is None:
        weights = model.get_weights()
    weights = [np.random.permutation(w.flat).reshape(w.shape) for w in weights]
    # Faster, but less random: only permutes along the first dimension
    # weights = [np.random.permutation(w) for w in weights]
    model.set_weights(weights)

def train():
        model.compile(loss='binary_crossentropy', optimizer='adadelta',metrics=['accuracy'])
        model.fit(x_train, y_train,
                        batch_size=16,
                        epochs=100,
                        verbose=1,
                        validation_data=(x_test, y_test))
        score = model.evaluate(x_test, y_test)
        # print(f'Score:{score}')
        res = model.predict(x_test)
        _res = np.zeros(y_test.shape)

        for k in range(res.shape[0]):
            _res[k][np.argmax(res[k])] = 1
        
        tp,fp,tn,fn = (0,0,0,0)
        for k in range (_res.shape[0]):
                if not (y_test[k]-_res[k]).all():
                        if np.argmax(y_test[k]) == 1:
                                tp+=1
                        else:
                                tn+=1
                else:
                        if np.argmax(y_test[k]) == 1:
                                fp+=1
                        else:
                                fn+=1
        global TP,TN,FP,FN
        TP+=tp
        TN+=tn
        FP+=fp
        FN+=fn

print("Training...")
for _ in range(1):
        train()               
print(f'|{TP/(TP+FP)}|{FP/(TP+FP)}|')
print('--------------------------------')
print(f'|{TN/(TN+FN)}|{FN/(TN+FN)}|')
